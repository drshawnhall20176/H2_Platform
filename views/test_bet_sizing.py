"""Offline tests for bet_sizing.py — sizing discipline and the xlsx snapshot."""

import io
import pandas as pd

import bet_sizing as BS


def _edges():
    # Two games. Game A is heavily favored/loaded (correlated), Game B is one modest play.
    return pd.DataFrame([
        {"Player": "A1", "Market": "Batter HR", "Side": "Over", "Line": 0.5, "Proj": 0.3,
         "ModelProb": 0.34, "Price": 300, "EV%": 8.0, "Game": "GameA"},
        {"Player": "A2", "Market": "Batter Total Bases", "Side": "Over", "Line": 1.5, "Proj": 1.6,
         "ModelProb": 0.62, "Price": -130, "EV%": 6.0, "Game": "GameA"},
        {"Player": "A3", "Market": "Batter Total Hits", "Side": "Over", "Line": 0.5, "Proj": 1.1,
         "ModelProb": 0.72, "Price": -200, "EV%": 5.0, "Game": "GameA"},
        {"Player": "B1", "Market": "Pitcher Strikeouts", "Side": "Over", "Line": 5.5, "Proj": 6.4,
         "ModelProb": 0.60, "Price": 110, "EV%": 6.0, "Game": "GameB"},
    ])


def test_shade_prob():
    assert BS.shade_prob(0.80, 5) == 0.75
    assert BS.shade_prob(0.03, 5) == 0.0        # floored, never negative
    assert BS.shade_prob(None, 5) == 0.0
    print("✓ shade_prob haircuts and floors correctly")


def test_odds_band():
    assert BS.in_odds_band(-200, -300, 400) is True
    assert BS.in_odds_band(-476, -300, 400) is False   # too juiced (heavy fav)
    assert BS.in_odds_band(+500, -300, 400) is False   # too long a shot
    assert BS.in_odds_band(+150, -300, 400) is True
    assert BS.in_odds_band("x", -300, 400) is False
    print("✓ odds band rejects heavy juice and long shots, keeps the middle")


def test_shading_can_zero_a_thin_edge():
    # A tiny raw edge should size to ~0 after a prudent haircut.
    df = pd.DataFrame([{"Player": "X", "ModelProb": 0.66, "Price": -181, "Game": "G"}])
    raw = BS.apply_stake_discipline(df, 150, shade_pts=0, per_game_pct=1.0)
    shaded = BS.apply_stake_discipline(df, 150, shade_pts=5, per_game_pct=1.0)
    assert raw["Stake $"].iloc[0] >= shaded["Stake $"].iloc[0]
    assert shaded["Stake $"].iloc[0] == 0.0
    print("✓ shading collapses a thin (66% vs ~64% market) edge to $0")


def test_per_game_cap():
    bankroll, cap = 150.0, 0.10          # $15 max per game
    out = BS.apply_stake_discipline(_edges(), bankroll, shade_pts=3,
                                    kelly_frac=0.25, cap_pct=0.05, per_game_pct=cap)
    gt = BS.game_totals(out)
    for _, row in gt.iterrows():
        assert row["Staked $"] <= bankroll * cap + 0.01, row
    # Per-bet cap (5%) also holds for every row.
    assert (out["Stake $"] <= bankroll * 0.05 + 0.01).all()
    print(f"✓ per-game cap holds (no game over ${bankroll*cap:.0f}); per-bet cap holds")


def test_no_cap_when_disabled():
    out = BS.apply_stake_discipline(_edges(), 150, per_game_pct=1.0)
    assert "Stake $" in out.columns and "Stake %" in out.columns
    print("✓ per_game_pct=1.0 disables the game cap without error")


def test_empty_frame():
    out = BS.apply_stake_discipline(pd.DataFrame(), 150)
    assert list(out.columns)[-3:] == ["Shaded %", "Stake $", "Stake %"]
    assert out.empty
    print("✓ empty edges frame is handled")


def test_build_card_xlsx():
    out = BS.apply_stake_discipline(_edges(), 150, shade_pts=3)
    out = out[out["Stake $"] > 0]
    card = out.rename(columns={"Price": "Odds"})[
        ["Player", "Market", "Side", "Line", "Proj", "Model %"
         if "Model %" in out.columns else "ModelProb", "Odds", "EV%",
         "Shaded %", "Stake $", "Stake %"]
    ].rename(columns={"ModelProb": "Model %"})
    data = BS.build_card_xlsx(card, bankroll=150.0, date_str="2026-07-01",
                              generated_at="2026-07-01 17:30", total_stake=float(card["Stake $"].sum()),
                              n_bets=len(card))
    assert data is not None and len(data) > 0
    # Re-open to confirm it's a valid workbook with our title.
    from openpyxl import load_workbook
    wb = load_workbook(io.BytesIO(data))
    ws = wb.active
    assert ws["A1"].value.startswith("H2 Sports")
    assert ws["A5"].value == "Player"       # header row
    print(f"✓ xlsx snapshot builds and re-opens ({len(data)} bytes, {len(card)} rows)")


if __name__ == "__main__":
    test_shade_prob()
    test_odds_band()
    test_shading_can_zero_a_thin_edge()
    test_per_game_cap()
    test_no_cap_when_disabled()
    test_empty_frame()
    test_build_card_xlsx()
    print("\nAll bet_sizing tests passed.")
