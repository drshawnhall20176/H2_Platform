"""
bet_sizing.py — the discipline layer for the Edge Board, factored out so it can be
unit-tested without Streamlit.
 
Ports the columns proven in the spreadsheet workflow:
  * Shaded %      — haircut the model probability before sizing (models are overconfident).
  * Stake $       — fractional Kelly on the SHADED probability, capped per bet.
  * Per-game cap  — scale every bet in a game down so one game can't dominate the roll
                    (props in the same game are correlated; independent Kelly overbets them).
  * Odds band     — keep prices inside a sane window (skip heavy juice and long shots).
 
Plus build_card_xlsx(): a formatted, static snapshot of tonight's sized card for sharing
(no formulas — it's a record of a decision, not a live model — so no weblink is exposed).
"""
 
from __future__ import annotations
 
from io import BytesIO
from typing import Optional
 
import pandas as pd
 
import odds_api as O
 
 
# --------------------------------------------------------------------------- sizing
def shade_prob(p, pts: float) -> float:
    """Model probability minus a flat honesty haircut, floored at 0. `pts` is in
    percentage POINTS (5 -> subtract 0.05), matching the spreadsheet's Shaded % column."""
    try:
        p = float(p)
    except (TypeError, ValueError):
        return 0.0
    return max(0.0, p - pts / 100.0)
 
 
def in_odds_band(american, floor: float, ceiling: float) -> bool:
    """True if an American price is inside [floor, ceiling]. floor is negative (heavy-fav
    guard, e.g. -300), ceiling positive (long-shot guard, e.g. +400). American odds are
    monotonic as raw numbers within this window, so a plain numeric compare is correct:
    -476 < -300 -> rejected; +500 > +400 -> rejected; -150 and +150 both pass."""
    try:
        a = float(american)
    except (TypeError, ValueError):
        return False
    return floor <= a <= ceiling
 
 
def apply_stake_discipline(edf: pd.DataFrame, bankroll: float, *, shade_pts: float = 5.0,
                           kelly_frac: float = 0.25, cap_pct: float = 0.05,
                           per_game_pct: float = 0.10) -> pd.DataFrame:
    """Add Shaded %, Stake $, Stake % to an edges frame.
 
    Sizing uses the SHADED probability, so a row with a small raw edge can legitimately
    size to $0 once you bet prudently — that's the point. Requires columns 'ModelProb'
    and 'Price'; uses 'Game' for the per-game cap if present."""
    edf = edf.copy()
    if edf.empty:
        for col in ("Shaded %", "Stake $", "Stake %"):
            edf[col] = pd.Series(dtype=float)
        return edf
 
    edf["Shaded %"] = edf["ModelProb"].map(lambda p: shade_prob(p, shade_pts))
    edf["Stake $"] = edf.apply(
        lambda r: O.kelly_stake(r["Shaded %"], r["Price"], bankroll, kelly_frac, cap_pct),
        axis=1,
    )
 
    # Per-game cap: props in one game rise and fall together, so Kelly (which assumes
    # independence) overbets the game. Scale each game's rows to keep its total <= cap.
    if "Game" in edf.columns and 0.0 < per_game_pct < 1.0 and bankroll > 0:
        game_cap = per_game_pct * bankroll
        totals = edf.groupby("Game")["Stake $"].transform("sum")
        factor = pd.Series(1.0, index=edf.index)
        over = totals > game_cap
        factor[over] = game_cap / totals[over]
        edf["Stake $"] = (edf["Stake $"] * factor).round(2)
 
    edf["Stake %"] = (edf["Stake $"] / bankroll) if bankroll else 0.0
    return edf
 
 
def game_totals(edf: pd.DataFrame) -> pd.DataFrame:
    """Post-cap dollars staked per game — handy for a 'no game over X%' sanity display."""
    if edf.empty or "Game" not in edf.columns or "Stake $" not in edf.columns:
        return pd.DataFrame(columns=["Game", "Staked $"])
    g = (edf[edf["Stake $"] > 0].groupby("Game")["Stake $"].sum()
         .reset_index().rename(columns={"Stake $": "Staked $"}))
    return g.sort_values("Staked $", ascending=False)
 
 
def stake_tier(stake, bankroll, dust_pct: float = 0.005) -> str:
    """Label a sized stake so the board reads at a glance:
      'No bet' — the edge shaded out to $0 (a hard, precise category).
      'Dust'   — positive but a negligible fraction of bankroll (below dust_pct); real but not
                 worth the click once you account for correlation and estimate noise.
      'Bet'    — enough edge survived shading that fractional-Kelly wants real money on it.
    Thresholds are a PERCENT of bankroll (not dollars), so the label stays honest at any roll
    size — $0.29 is dust on $100 but a real bet on $10,000, and this scales with it."""
    try:
        s, bk = float(stake), float(bankroll)
    except (TypeError, ValueError):
        return ""
    if bk <= 0 or s <= 0:
        return "No bet"
    return "Dust" if (s / bk) < dust_pct else "Bet"
 
 
# --------------------------------------------------------------------------- export
_MONEY = {"Stake $"}
_PCT_FRAC = {"Model %", "Shaded %", "Stake %"}   # stored as fractions (0.75 -> 75.0%)
_EV_UNITS = {"EV%"}                               # stored already in percent units (5.3)
_FMT = {"Line": "0.0", "Proj": "0.00", "Odds": "0"}
 
 
def build_card_xlsx(card_df: pd.DataFrame, *, bankroll: float, date_str: str,
                    generated_at: str, total_stake: float, n_bets: int) -> Optional[bytes]:
    """A formatted, self-contained .xlsx snapshot of tonight's sized card for sharing.
 
    Static values only (no formulas, no links) — it's a record of a decision, so it never
    references your model or a URL. Returns bytes, or None if openpyxl isn't available."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
    except Exception:
        return None
 
    wb = Workbook()
    ws = wb.active
    ws.title = "Tonight's Card"
 
    ws["A1"] = "H2 Sports — Tonight's Sized Card"
    ws["A1"].font = Font(name="Arial", size=14, bold=True)
    ws["A2"] = (f"Slate {date_str}   ·   generated {generated_at}   ·   "
                f"bankroll ${bankroll:,.2f}   ·   {n_bets} bet(s)   ·   ${total_stake:,.2f} staked")
    ws["A2"].font = Font(name="Arial", size=10, italic=True, color="555555")
    ws["A3"] = ("Selections for analysis and entertainment — not a guarantee. "
                "Stakes are fractional-Kelly, shaded and capped. Bet responsibly.")
    ws["A3"].font = Font(name="Arial", size=9, italic=True, color="888888")
 
    headers = list(card_df.columns)
    hdr_row = 5
    for j, h in enumerate(headers, start=1):
        c = ws.cell(row=hdr_row, column=j, value=h)
        c.font = Font(name="Arial", bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", start_color="2E7D32")
        c.alignment = Alignment(horizontal="center")
 
    for i, (_, row) in enumerate(card_df.iterrows(), start=hdr_row + 1):
        for j, h in enumerate(headers, start=1):
            v = row[h]
            if pd.isna(v):
                v = None
            elif hasattr(v, "item"):      # numpy scalar -> native python
                v = v.item()
            ws.cell(row=i, column=j, value=v)
 
    n = len(card_df)
    for j, h in enumerate(headers, start=1):
        col = get_column_letter(j)
        fmt = None
        if h in _MONEY:
            fmt = '$#,##0.00;($#,##0.00);"-"'
        elif h in _PCT_FRAC:
            fmt = "0.0%"
        elif h in _EV_UNITS:
            fmt = '+0.0"%";-0.0"%";0"%"'
        elif h in _FMT:
            fmt = _FMT[h]
        widths = [len(str(h))] + [len(str(card_df.iloc[k][h])) for k in range(n)]
        ws.column_dimensions[col].width = min(max(max(widths) + 2, 9), 42)
        for r in range(hdr_row + 1, hdr_row + 1 + n):
            cell = ws.cell(row=r, column=j)
            cell.font = Font(name="Arial")
            if fmt:
                cell.number_format = fmt
 
    ws.freeze_panes = f"A{hdr_row + 1}"
 
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
